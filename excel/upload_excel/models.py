import re
import string
import uuid
from typing import Any, Callable, List, Optional, Tuple, TypeVar, Union

import numpy as np
import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import models
from django.db.utils import ProgrammingError
from django.http import HttpRequest
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from upload_excel.utils.cell_tree import CellNode, CellTree
from upload_excel.utils.sort import A2ZListMaker

_T = TypeVar("_T", bound=models.Model)
_F = TypeVar("_F", bound=models.Field)
_CRM = TypeVar("_CRM", bound="CellRangeModel")
_ESM = TypeVar("_ESM", bound="ExcelSheetModel")
_EST = TypeVar("_EST", bound="ESTemplateNamesModel")
_CM = TypeVar("_CM", bound="ColumnModel")
_RM = TypeVar("_RM", bound="RowModel")
_CTM = TypeVar("_CTM", bound="ContentModel")

# 改行パターン
br_pattern: str = "?#$%&@!?*+"


getters = {
    "digit": re.compile(r"\d+").findall,
    "alphabet": re.compile(r"([A-Z]+)").findall
}


def get_bound_items(cell_range: CellRange,
                    bound_type: str = "digit") -> Tuple[str, str]:
    # preparations
    if bound_type not in getters:
        raise KeyError("'bound_type' must be either ''digit' or 'alphabet'")

    getter: str = getters[bound_type]
    str_coord: str = cell_range.coord

    if ":" not in str_coord:
        str_coord = ":".join([str_coord, str_coord])

    # set
    start: str
    end: str
    start, end = str_coord.split(":")
    start = getter(start)[0]
    end = getter(end)[0]

    return start, end


def get_cell_value(cell: Cell, concat_size: int = 0) -> Optional[str]:
    val: Optional[Any] = cell.value
    output: str = ""
    if val is not None:
        if concat_size > 0:
            output += br_pattern
        output += str(val)
    return output


def get_text(cell):
    val = cell.value
    output = ""
    if val is not None:
        output += str(val)
    return output


class ESTemplateNamesModel(models.Model):
    name: _F = models.CharField(
        verbose_name="シートタイプ",
        blank=False,
        null=False,
        default="profile",
        editable=True,
        max_length=20,
    )
    init_choices: List[str] = [
        (1, "profile"), (2, "project")
    ]
    class Meta:
        db_table: str = "template_names"

    @classmethod
    def add(cls, val: str) -> None:
        template_name_model: _EST = cls(name=val)
        template_name_model.save(force_insert=True)

    @classmethod
    def _get_template_names(cls, *args, **kwargs) -> List[str]:
        output: List[str] = [] + cls.init_choices
        count: int = 3
        _choices: List[str] = [c for _, c in cls.init_choices]
        # exception process when using makemigration command as initialization.
        try:
            for v in ESTemplateNamesModel.objects.all().values_list():
                if v in _choices:
                    continue
                output += [(count, v)]
                count += 1
                _choices += [v]
            return output
        except ProgrammingError:
            return output

class ExcelSheetModel(models.Model):
    sheet_create_time: _F = models.DateTimeField(
        verbose_name="シート作成日時",
        blank=False,
        null=False,
        default=timezone.now,
        help_text=(
            "Define when a sheet has created."
        ),
    )
    sheet_update_time: _F = models.DateTimeField(
        verbose_name="シート作成日時",
        blank=False,
        null=False,
        default=timezone.now,
        help_text=(
            "Define when a sheet has updated."
        )
    )
    sheet_id: _F = models.UUIDField(
        verbose_name="シートID",
        primary_key=True,
        blank=False,
        null=False,
        default=uuid.uuid4(),
        editable=True,
        help_text=(
            "An ID based on uuid4 is assigned by sheet randomly and automatically."
        )
    )
    sheet_type: _F = models.CharField(
        verbose_name="シートタイプ",
        blank=False,
        null=False,
        default="profile",
        editable=True,
        max_length=10,
        help_text=(
            "Set a name of template sheet."
            "'profile' that is default sheet type is a template "
            "we designed as typical profile in our office"
            "Even before you don't generate your custom template, "
            "you can use two template types of 'profile' and 'project'"
        ),
        choices=ESTemplateNamesModel._get_template_names()
    )
    col_size: _F = models.PositiveIntegerField(
        verbose_name="シートのカラムサイズ",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    row_size: _F = models.PositiveIntegerField(
        verbose_name="シートのローサイズ",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    excel_matrix: np.ndarray
    child_rate: float = 0.95
    class Meta:
        db_table: str = "excel_sheet"

    @classmethod
    def is_valid_request(self,
                         request: HttpRequest,
                         file_key: str = "file"
                         ) -> None:
        is_valid: bool = file_key in request.FILES
        if not is_valid:
            raise KeyError()

    @classmethod
    def get_binary_data(cls,
                        request: HttpRequest,
                        file_key: str = "file") -> str:
        file_data: List[InMemoryUploadedFile] = request.FILES[file_key]
        return file_data.file

    @classmethod
    def create_model(cls,
                     request: HttpRequest,
                     file_key: str = "file",
                     sheet_type: str = "profile") -> _ESM:
        cls.is_valid_request(request, file_key)
        binary: str = cls.get_binary_data(request, file_key)

        workbook: Workbook = openpyxl.load_workbook(binary)
        worksheet: Worksheet = workbook.active

        excel_sheet_model: _ESM = cls(sheet_id=uuid.uuid4(),
                                      sheet_type=sheet_type,
                                      col_size=worksheet.max_column,
                                      row_size=worksheet.max_row)
        excel_sheet_model.excel_matrix =\
            np.zeros((100, worksheet.max_column))
        excel_sheet_model.save(force_update=True)

        excel_sheet_model.create_cell_ranges(worksheet)

        workbook.close()
        return excel_sheet_model

    def create_cell_ranges(self, worksheet: Worksheet) -> None:
        # TODO メソッドを細かく分ける
        list_maker = A2ZListMaker(max_size=worksheet.max_column, init=string.ascii_uppercase)
        list_maker.create()

        cell_ranges: List[CellRange] = worksheet.merged_cells
        excel_array = np.zeros((worksheet.max_row, worksheet.max_column))
        ranges: CellRange
        idx: int

        coord_list: List[str] = []
        for idx, ranges in enumerate(cell_ranges):
            coords = ranges.coord.split(":")

            out = []
            for c in coords:
                out += getters["alphabet"](c)
                out += getters["digit"](c)
            coord_list.append(out)

        out_map = {}
        rect: dict[str, int] = {
            "upper_left": 10000,
            "upper_right": 0,
            "lower_left": 10000,
            "lower_right": 0
        }
        for n, out in enumerate(coord_list):
            cs, rs, ce, re = out
            txt = ""
            for cells in worksheet[cs+rs+":"+ce+re]:
                for cell in cells:
                    txt += get_text(cell)

            if len(txt) < 1:
                continue

            cs = list_maker.values.index(cs)
            ce = list_maker.values.index(ce) + 1
            rs = int(rs) - 1
            re = int(re)

            rect["upper_left"] = min([rect["upper_left"], rs])
            rect["upper_right"] = max([rect["upper_right"], re])
            rect["lower_left"] = min([rect["lower_left"], cs])
            rect["lower_right"] = max([rect["lower_right"], ce])

            # TODO null部分を縦1x横上に合わせて最大の長方形として全て定義し直す。値は空。
            array_mask = excel_array[rs:re, cs:ce] == 0
            excel_array[rs:re, cs:ce][array_mask] = n + 1
            out_map[n + 1] = {
                "text": txt, "ranges": [(cs, ce), (rs, re)],
                "merged_cell": cell_ranges.ranges[n]
            }

        excel_mask = np.ones_like(excel_array, dtype=bool)
        excel_mask[rect["upper_left"]:rect["upper_right"], rect["lower_left"]:rect["lower_right"]] = False
        excel_array[excel_mask] = None
        count = int(np.nanmax(excel_array)) + 1

        for row_idx, row in enumerate(excel_array):
            width_list = []
            null_row = row == 0
            if not np.any(null_row):
                continue

            if null_row[0]:
                init = np.ones
            else:
                init = np.zeros

            shift_null = np.r_[init((1, ), dtype=bool), null_row[:-1]]

            start_points = (null_row & ~shift_null).nonzero()[0]
            end_points = (~null_row & shift_null).nonzero()[0]

            if null_row[0]:
                start_points = np.r_[np.array([0]), start_points]

            if len(start_points) > len(end_points):
                end_points = np.r_[end_points, np.array([None])]

            for s, e in zip(start_points, end_points):
                width_list.append(s)
                width_list.append(e)

            for i in range(len(width_list)//2):
                start = width_list[2 * i]
                end = width_list[2 * i + 1]

                if end is not None:
                    if np.any(excel_array[row_idx, start:end] > 0):
                        raise IndexError()
                    excel_array[row_idx, start:end] = count
                else:
                    if np.any(excel_array[row_idx, start:] > 0):
                        raise IndexError()
                    excel_array[row_idx, start:] = count

                if end is None:
                    end = len(row) - 1

                start_cell = list_maker.values[start] + str(row_idx + 1)
                end_cell = list_maker.values[end - 1] + str(row_idx + 1)
                merged_cell = start_cell + ":" + end_cell
                merged_cell = CellRange(range_string=merged_cell)

                out_map[count] = {
                    "text": "", "ranges": [(start, end + 1), (row_idx, row_idx + 1)],
                    "merged_cell": merged_cell
                }
                count += 1

        if np.nansum(excel_array == 0) > 0:
            raise ValueError(
                "No Zero must be included, "
                f"but there is {np.nansum(excel_array == 0)} zeros in 'excel_array'")

        cell_content: dict[int, str] = {
            idx: contents["text"] for idx, contents in out_map.items()
        }
        tree = CellTree.create_tree(excel_array,
                                    child_rate=self.child_rate,
                                    cell_content=cell_content)

        for idx, outs in out_map.items():
            CellRangeModel.\
                create_model(self,
                             worksheet=worksheet,
                             cell_range=outs["merged_cell"],
                             idx=idx,
                             node=tree.tree[idx])

class CellRangeModel(models.Model):
    excel_sheet: _F = models.ForeignKey(
        ExcelSheetModel,
        on_delete=models.CASCADE,
        related_name="cell_ranges"
    )
    sheet_create_time: _F = models.DateTimeField(
        verbose_name="セル範囲作成日時",
        blank=False,
        null=False,
        default=timezone.now,
        help_text=(
            "Define when a sheet has created."
        ),
    )
    sheet_update_time: _F = models.DateTimeField(
        verbose_name="セル範囲作成日時",
        blank=False,
        null=False,
        default=timezone.now,
        help_text=(
            "Define when a sheet has updated."
        )
    )
    cell_range_id: _F = models.UUIDField(
        verbose_name="セル範囲ID",
        blank=False,
        null=False,
        default=uuid.uuid4(),
        editable=True,
        help_text=(
            "An ID based on uuid4 is assigned by cell range randomly and automatically."
        )
    )
    cell_range_id_by_order: _F = models.IntegerField(
        verbose_name="順序付きセル範囲ID",
        blank=False,
        null=False,
        default=None,
        editable=True,
        help_text=(
            "An ID based on merged order is assigned by cell range."
        )
    )
    effective_cell_width: _F = models.IntegerField(
        verbose_name="有効セル幅",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    effective_cell_height: _F = models.IntegerField(
        verbose_name="有効セル高さ",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    has_parent: _F = models.BooleanField(
        verbose_name="親の有無",
        blank=False,
        null=False,
        default=True,
        editable=True,
    )
    is_dev_exp_id: _F = models.BooleanField(
        verbose_name="開発経験IDかどうか",
        blank=False,
        null=False,
        default=False,
        editable=True,
    )
    class Meta:
        db_table: str = "cell_range"

    @classmethod
    def create_model(cls,
                     excel_sheet: _ESM,
                     worksheet: Worksheet,
                     cell_range: CellRange,
                     idx: int = 0,
                     node: CellNode = CellNode()) -> _CRM:

        # When creating a child model,
        # an inputting parent model which has been defined
        # as foreign key model at child must be saved before.
        crm: _CRM = cls(excel_sheet=excel_sheet,
                        cell_range_id_by_order=idx,
                        effective_cell_width=node.width,
                        effective_cell_height=node.height,
                        has_parent=node.has_parent(),
                        is_dev_exp_id=node.is_dev_experience())
        crm.save(force_insert=True)
        ColumnModel.create_model(cell_range_model=crm,
                                 cell_range=cell_range,
                                 idx=idx)
        RowModel.create_model(cell_range_model=crm,
                              cell_range=cell_range,
                              idx=idx)
        ContentModel.create_model(cell_range_model=crm,
                                  worksheet=worksheet,
                                  cell_range=cell_range,
                                  idx=idx)
        return crm


class ColumnModel(models.Model):
    cell_range: _F = models.ForeignKey(
        CellRangeModel,
        on_delete=models.CASCADE,
        related_name="columns"
    )
    cell_start: _F = models.CharField(
        verbose_name="セルのカラム初期位置",
        blank=False,
        null=False,
        default="A1",
        editable=True,
        max_length=10,
    )
    cell_end: _F = models.CharField(
        verbose_name="セルのカラム終端位置",
        blank=False,
        null=False,
        default="A1",
        editable=True,
        max_length=10,
    )
    cell_size: _F = models.PositiveIntegerField(
        verbose_name="セルの列方向サイズ",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    cell_range_id_by_order: _F = models.IntegerField(
        verbose_name="順序付きセル範囲ID",
        blank=False,
        null=False,
        default=None,
        editable=True,
        help_text=(
            "An ID based on merged order is assigned by cell range."
        )
    )

    class Meta:
        db_table: str = "column"

    @classmethod
    def create_model(cls,
                     cell_range_model: _CRM,
                     cell_range: CellRange,
                     idx: int = 0) -> _CM:
        start, end = get_bound_items(cell_range, bound_type="alphabet")
        cell_size: int = cell_range.max_col - cell_range.min_col + 1
        column: _CM = cls(cell_range=cell_range_model,
                         cell_start=start,
                         cell_end=end,
                         cell_size=cell_size,
                         cell_range_id_by_order=idx)
        column.save(force_insert=True)
        return column


class RowModel(models.Model):
    cell_range: _F = models.ForeignKey(
        CellRangeModel,
        on_delete=models.CASCADE,
        related_name="rows"
    )
    cell_start: _F = models.CharField(
        verbose_name="セルのロー初期位置",
        blank=False,
        null=False,
        default="A1",
        editable=True,
        max_length=10,
    )
    cell_end: _F = models.CharField(
        verbose_name="セルのロー終端位置",
        blank=False,
        null=False,
        default="A1",
        editable=True,
        max_length=10,
    )
    cell_size: _F = models.PositiveIntegerField(
        verbose_name="セルの行方向サイズ",
        blank=False,
        null=False,
        default=1,
        editable=True,
    )
    cell_range_id_by_order: _F = models.IntegerField(
        verbose_name="順序付きセル範囲ID",
        blank=False,
        null=False,
        default=None,
        editable=True,
        help_text=(
            "An ID based on merged order is assigned by cell range."
        )
    )

    class Meta:
        db_table: str = "row"

    @classmethod
    def create_model(cls,
                     cell_range_model: _CRM,
                     cell_range: CellRange,
                     idx: int = 0) -> _RM:
        start, end = get_bound_items(cell_range, bound_type="digit")
        cell_size: int = cell_range.max_row - cell_range.min_row + 1
        row: _RM = cls(cell_range=cell_range_model,
                       cell_start=start,
                       cell_end=end,
                       cell_size=cell_size,
                       cell_range_id_by_order=idx)
        row.save(force_insert=True)
        return row



class ContentModel(models.Model):
    cell_range: _F = models.ForeignKey(
        CellRangeModel,
        on_delete=models.CASCADE,
        related_name="content"
    )
    cell_content: _F = models.TextField(
        verbose_name="セルの内容",
        blank=False,
        null=False,
        default="",
        editable=True,
    )
    cell_range_id_by_order: _F = models.IntegerField(
        verbose_name="順序付きセル範囲ID",
        blank=False,
        null=False,
        default=None,
        editable=True,
        help_text=(
            "An ID based on merged order is assigned by cell range."
        )
    )
    class Meta:
        db_table: str = "content"

    @classmethod
    def extract_cell_content(cls,
                             worksheet: Worksheet,
                             cell_range: CellRange) -> str:
        coord: str = cell_range.coord
        if ":" not in coord:
            coord = ":".join([coord, coord])

        start, end = coord.split(":")
        cell_info: Tuple[Tuple[Cell, ...]] = worksheet[start:end]
        output: str = ""
        for cell_row in cell_info:
            for cell in cell_row:
                concat_size = len(output)
                output += get_cell_value(cell, concat_size)
        return output

    @classmethod
    def create_model(cls,
                     cell_range_model: _CRM,
                     worksheet: Worksheet,
                     cell_range: CellRange,
                     idx: int = 0) -> _CTM:
        cell_content = cls.extract_cell_content(worksheet, cell_range)
        content: _CTM = cls(cell_range=cell_range_model,
                            cell_content=cell_content,
                            cell_range_id_by_order=idx)
        content.save(force_insert=True)
        return content
